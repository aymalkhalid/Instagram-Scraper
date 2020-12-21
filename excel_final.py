# import openpyxl module 
import openpyxl 
  
# Call a Workbook() function of openpyxl  
# to create a new blank Workbook object 
wb = openpyxl.Workbook() 
  
# Get workbook active sheet   
# from the active attribute 
sheet = wb.active 
  
# Cell objects also have row, column 
# and coordinate attributes that provide 
# location information for the cell.
c3 = sheet['A2'] 
c3.value = "RAHUL"

userinfo={'username': 'aymalkhalid', 'user id': '3540216858', 'name': 'Aymal Khalid', 'followers': 553, 'following': 854, 'posts img': 129, 'posts vid': 0, 'reels': 2, 'bio': 'Nothing in life is permanent.', 'external url': None, 'private': False, 'verified': False, 'profile img': 'https://tinyurl.com/yaeaj2r6', 'business account': False, 'joined recently': False, 'business category': None, 'category': None, 'has guides': False}
#Writing Keys to Excel
global columns
columns=[]
for key in userinfo:
   columns.append(key)
column_count=1
column_count =int(column_count)
for col in columns:
    c1 = sheet.cell(row = 1, column = column_count)    
    column_count=column_count+1
    c1.value= str(col)
#Writing Values to Excel
global values
values=[]
for key in userinfo:
   values.append(userinfo[key])
row_count=1
row_count=int(row_count)
for row in values:
    c2 = sheet.cell(row = 2, column = row_count) 
    row_count=row_count+1
    c2.value=str(row)
wb.save("D:\\Scrapper_Python\\1\\.lib\\demo.xlsx") 
